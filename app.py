"""
Streamlit Web App: FEHB + FEDVIP 2026 Cost Comparison Tool

This web application replicates the core functionality of the provided
Excel‑based FEHB + FEDVIP 2026 General Use Calculator.  Users can enter
their household details, utilisation assumptions and optional dental and
vision premiums to estimate total health care spending for every available
Federal Employee Health Benefit (FEHB) plan.  The application reads
plan data from the supplied workbook and applies a simplified model
derived from the workbook’s formulas and underlying assumptions.  Cost
estimates are broken down into premium and out‑of‑pocket components and
displayed alongside each plan’s name, network type and an inferred
coverage tier (Basic, Standard, Rich or High Deductible).

Key simplifying assumptions:

* Plan categories are inferred from the plan option name.  Plans
  containing “HDHP” are treated as High Deductible, those containing
  “High” or “Rich” as Rich, “Standard” as Standard and “Basic” as
  Basic.  If no keywords are present the plan is treated as Standard.
* Out‑of‑pocket costs for medical services are calculated by
  multiplying user‑specified visit counts by category specific copay
  values from the model settings sheet.  Surgery, therapy and
  maternity costs are modelled using the average allowed charges and
  plan pay percentages defined in the same sheet.
* Major dental work is estimated using the number of crowns or
  implants entered by the user.  Plan coverage is assumed to pay
  50 % of major dental work up to a per‑person annual maximum of
  $2,000.  Costs in excess of the coverage limit are borne by the
  enrollee.  Dental premiums can be entered manually on the input
  form.
* Vision plans are represented only by their premium; no vision
  out‑of‑pocket costs are modelled.
* HSA/FSA contributions reduce out‑of‑pocket costs but cannot lower
  total spending below the premium component.

The resulting tool provides a quick and transparent way for users to
explore how different plans might impact their annual healthcare
spending.  While it does not implement the full Excel calculator
logic, it leverages the same parameter values and plan premiums to
deliver a faithful and insightful approximation.
"""

import pandas as pd
import streamlit as st
from pathlib import Path
import openpyxl


@st.cache_data
def load_workbook_data(workbook_path: Path):
    """Load plan and model settings from the provided Excel workbook.

    The workbook is expected to contain at least two worksheets:

    * ``Plans`` – holds a row for each FEHB plan with columns for plan
      name, option, network type and annual premiums.
    * ``Model Settings`` – contains parameter/value pairs used to
      calculate out‑of‑pocket costs.

    Parameters
    ----------
    workbook_path : Path
        Path to the Excel workbook on disk.

    Returns
    -------
    tuple[pd.DataFrame, dict]
        A tuple where the first item is a DataFrame of plans and the
        second is a dictionary of model settings keyed by the
        parameter name.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    # Load plans sheet into a DataFrame
    plans_ws = wb["Plans"]
    # Extract header names
    headers = [cell.value for cell in plans_ws[1]]
    plans_data = []
    for row in plans_ws.iter_rows(min_row=2, values_only=True):
        plans_data.append(row)
    plans_df = pd.DataFrame(plans_data, columns=headers)
    # Only keep visible plans
    if "Visible?" in plans_df.columns:
        plans_df = plans_df[plans_df["Visible?"].astype(str).str.upper() == "Y"].copy()
    # Drop completely empty columns at the end
    plans_df = plans_df.loc[:, ~plans_df.columns.str.contains('^Unnamed')]  # drop Excel unnamed columns

    # Load model settings
    settings_ws = wb["Model Settings"]
    settings = {}
    for parameter, value, explanation in settings_ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if parameter is None:
            continue
        settings[str(parameter)] = value
    return plans_df, settings


def infer_plan_category(plan_name: str) -> str:
    """Infer plan coverage tier from the plan name.

    This helper examines the plan name for common keywords to map
    options into one of four tiers: ``HDHP``, ``RICH``, ``STANDARD`` or
    ``BASIC``.  If no keyword is found the plan defaults to
    ``STANDARD``.

    Parameters
    ----------
    plan_name : str
        The full plan option name.

    Returns
    -------
    str
        Inferred tier label.
    """
    name = plan_name.lower()
    if "hdhp" in name or "high deductible" in name or "consumer" in name:
        return "HDHP"
    if "high" in name or "rich" in name:
        return "RICH"
    if "standard" in name:
        return "STANDARD"
    if "basic" in name:
        return "BASIC"
    # default fallback
    return "STANDARD"


def compute_plan_cost(
    plan: pd.Series,
    settings: dict,
    family_size: int,
    annual_income: float,
    monthly_rx: int,
    pcp_visits: int,
    spec_visits: int,
    urgent_visits: int,
    major_surgery: bool,
    therapy_program: bool,
    maternity: bool,
    major_dental: bool,
    num_crowns: int,
    include_dental: bool,
    dental_premium_monthly: float,
    include_vision: bool,
    vision_premium_monthly: float,
    use_fsa_hsa: bool,
    contribution_amount: float,
) -> dict:
    """Calculate total cost for a single plan.

    The calculation approximates the Excel formula logic.  It
    decomposes costs into premium and out‑of‑pocket components and
    accounts for optional dental and vision coverage as well as an
    HSA/FSA contribution.  Returns a dictionary with the results.
    """
    # Determine plan tier
    category = infer_plan_category(plan["Plan Option Name"])
    # Pull copays and plan pay percentages from settings using the category
    rx_copay = settings.get(f"Rx copay {category}", 0)
    pcp_copay = settings.get(f"PCP copay {category}", 0)
    spec_copay = settings.get(f"Spec/PT copay {category}", 0)
    urgent_copay = settings.get(f"Urgent copay {category}", 0)
    # Base allowed charges for high‑cost events
    surgery_allowed = settings.get("Surgery average allowed", 0)
    therapy_allowed = settings.get("Therapy program cost (base)", 0)
    maternity_allowed = settings.get("Maternity episode allowed", 0)
    dental_crown_cost = settings.get("Crown unit cost", 0)
    dental_coverage = settings.get("Dental major coverage %", 0)  # e.g. 0.5
    dental_annual_max = settings.get("Dental annual max per person", 0)
    # Plan pay % for high cost events
    surgery_plan_pay = settings.get(f"Surgery plan pay % {category}", 0)
    therapy_plan_pay = settings.get(f"Therapy plan pay % {category}", 0)
    maternity_plan_pay = settings.get(f"Maternity plan pay % {category}", 0)
    # Calculate OOP components
    rx_cost = monthly_rx * 12 * rx_copay
    pcp_cost = pcp_visits * pcp_copay
    spec_cost = spec_visits * spec_copay
    urgent_cost = urgent_visits * urgent_copay
    surgery_cost = 0
    if major_surgery:
        # Patient pays the portion not covered by plan
        surgery_cost = (1 - surgery_plan_pay) * surgery_allowed
    therapy_cost = 0
    if therapy_program:
        therapy_cost = (1 - therapy_plan_pay) * therapy_allowed
    maternity_cost = 0
    if maternity:
        maternity_cost = (1 - maternity_plan_pay) * maternity_allowed
    # Dental costs
    dental_cost = 0
    dental_premium = 0
    if include_dental:
        dental_premium = dental_premium_monthly * 12
        if major_dental and num_crowns > 0:
            total_dental_procedure_cost = num_crowns * dental_crown_cost
            # Plan covers coverage% up to family max
            max_plan_pay = dental_annual_max * family_size
            plan_pay = min(total_dental_procedure_cost * dental_coverage, max_plan_pay)
            dental_cost = total_dental_procedure_cost - plan_pay
    # Vision costs
    vision_cost = 0
    vision_premium = 0
    if include_vision:
        vision_premium = vision_premium_monthly * 12
        # OOP for vision is not modelled separately
        vision_cost = 0
    # Compute total premium (medical + optional dental + optional vision)
    med_premium = plan.get("Medical_Premium_Annual", 0)
    total_premium = med_premium + dental_premium + vision_premium
    # Total OOP before FSA/HSA
    total_oop = rx_cost + pcp_cost + spec_cost + urgent_cost + surgery_cost + therapy_cost + maternity_cost + dental_cost + vision_cost
    # Apply FSA/HSA contribution: reduce OOP but not below zero
    if use_fsa_hsa:
        total_oop = max(total_oop - contribution_amount, 0)
    total_cost = total_premium + total_oop
    # Percent of income
    percent_income = (total_cost / annual_income) if annual_income > 0 else 0
    return {
        "Plan Option Name": plan["Plan Option Name"],
        "Plan Option Type": plan.get("Plan Option Type", ""),
        "Network Type": plan.get("Network Type", ""),
        "Plan Category": category,
        "Medical Premium": med_premium,
        "Dental Premium": dental_premium,
        "Vision Premium": vision_premium,
        "Total Premium": total_premium,
        "Rx Cost": rx_cost,
        "PCP Cost": pcp_cost,
        "Specialist/PT Cost": spec_cost,
        "Urgent Care Cost": urgent_cost,
        "Surgery Cost": surgery_cost,
        "Therapy Cost": therapy_cost,
        "Maternity Cost": maternity_cost,
        "Dental OOP Cost": dental_cost,
        "Vision OOP Cost": vision_cost,
        "Total OOP": total_oop,
        "Total Cost": total_cost,
        "% of Income": percent_income,
    }


def main():
    st.set_page_config(page_title="FEHB 2026 Cost Comparison Tool", layout="wide")
    st.title("FEHB + FEDVIP 2026 Cost Comparison Tool")
    st.markdown(
        """
        This interactive tool allows you to estimate your annual healthcare
        spending across all available Federal Employee Health Benefit (FEHB)
        plans.  Provide your household details below and the app will
        calculate premiums and out‑of‑pocket costs using a simplified
        version of the official 2026 calculator.  The calculations use
        plan premiums and model parameters extracted directly from the
        provided spreadsheet.
        """
    )
    # Load workbook
    workbook_path = Path("FEHB_2026_General_Use_Calculator_PRO_v1.7_Failsafe.xlsx")
    if not workbook_path.exists():
        st.error(
            f"Workbook {workbook_path.name} not found. Please ensure it is present in the same directory as this app."
        )
        return
    plans_df, settings = load_workbook_data(workbook_path)
    # Sidebar for inputs
    st.sidebar.header("Your Information")
    family_size = st.sidebar.number_input("Family size", min_value=1, max_value=10, value=2)
    annual_income = st.sidebar.number_input("Annual household income ($)", min_value=0.0, value=75000.0, step=1000.0)
    utilization_level = st.sidebar.selectbox("Utilization level", ["Low", "Moderate", "High"], index=1)
    # Suggest default visit counts based on utilization level
    default_settings = {
        "Low": {"PCP": 2, "Spec": 4, "Urgent": 1, "RX": 1},
        "Moderate": {"PCP": 6, "Spec": 12, "Urgent": 2, "RX": 10},
        "High": {"PCP": 10, "Spec": 24, "Urgent": 4, "RX": 20},
    }
    defaults = default_settings[utilization_level]
    monthly_rx = st.sidebar.number_input("Monthly prescription count", min_value=0, max_value=60, value=defaults["RX"])
    pcp_visits = st.sidebar.number_input("Annual primary care visits", min_value=0, max_value=50, value=defaults["PCP"])
    spec_visits = st.sidebar.number_input("Annual specialist/therapy visits", min_value=0, max_value=50, value=defaults["Spec"])
    urgent_visits = st.sidebar.number_input("Annual urgent care visits", min_value=0, max_value=20, value=defaults["Urgent"])
    st.sidebar.header("High‑Cost Events")
    major_surgery = st.sidebar.checkbox("Major surgery this year?", value=False)
    therapy_program = st.sidebar.checkbox("Therapy program?", value=False)
    maternity = st.sidebar.checkbox("Maternity event?", value=False)
    st.sidebar.header("Dental & Vision")
    include_dental = st.sidebar.checkbox("Include dental coverage?", value=True)
    num_crowns = 0
    dental_premium_monthly = 0.0
    major_dental = False
    if include_dental:
        major_dental = st.sidebar.checkbox("Major dental work?", value=False, help="Crowns or implants this year")
        if major_dental:
            num_crowns = st.sidebar.number_input("Number of crowns/implants", min_value=1, max_value=10, value=1)
        dental_premium_monthly = st.sidebar.number_input(
            "Dental premium ($/month)", min_value=0.0, max_value=1000.0, value=50.0, step=1.0
        )
    include_vision = st.sidebar.checkbox("Include vision coverage?", value=True)
    vision_premium_monthly = 0.0
    if include_vision:
        vision_premium_monthly = st.sidebar.number_input(
            "Vision premium ($/month)", min_value=0.0, max_value=500.0, value=10.0, step=1.0
        )
    st.sidebar.header("HSA/FSA")
    use_fsa_hsa = st.sidebar.checkbox("Use HSA/FSA contribution?", value=False)
    contribution_amount = 0.0
    if use_fsa_hsa:
        contribution_amount = st.sidebar.number_input(
            "Annual HSA/FSA contribution ($)", min_value=0.0, max_value=10000.0, value=2000.0, step=100.0
        )
    # Filter options: allow user to limit by network type
    st.sidebar.header("Plan Filters")
    network_options = sorted(plans_df["Network Type"].dropna().unique().tolist())
    selected_networks = st.sidebar.multiselect(
        "Network types (leave blank for all)", options=network_options, default=network_options
    )
    # When user presses button, compute costs
    st.header("Estimated Costs by Plan")
    if st.button("Run comparison"):
        # Filter by network types if selected
        filtered_plans = plans_df.copy()
        if selected_networks:
            filtered_plans = filtered_plans[filtered_plans["Network Type"].isin(selected_networks)]
        results = []
        for _, plan in filtered_plans.iterrows():
            res = compute_plan_cost(
                plan,
                settings,
                family_size=family_size,
                annual_income=annual_income,
                monthly_rx=monthly_rx,
                pcp_visits=pcp_visits,
                spec_visits=spec_visits,
                urgent_visits=urgent_visits,
                major_surgery=major_surgery,
                therapy_program=therapy_program,
                maternity=maternity,
                major_dental=major_dental,
                num_crowns=int(num_crowns),
                include_dental=include_dental,
                dental_premium_monthly=dental_premium_monthly,
                include_vision=include_vision,
                vision_premium_monthly=vision_premium_monthly,
                use_fsa_hsa=use_fsa_hsa,
                contribution_amount=contribution_amount,
            )
            results.append(res)
        results_df = pd.DataFrame(results)
        # Sort by total cost ascending
        results_df = results_df.sort_values("Total Cost")
        # Format currency columns
        currency_cols = [
            "Medical Premium",
            "Dental Premium",
            "Vision Premium",
            "Total Premium",
            "Rx Cost",
            "PCP Cost",
            "Specialist/PT Cost",
            "Urgent Care Cost",
            "Surgery Cost",
            "Therapy Cost",
            "Maternity Cost",
            "Dental OOP Cost",
            "Vision OOP Cost",
            "Total OOP",
            "Total Cost",
        ]
        percent_col = "% of Income"
        def format_currency(x):
            return f"${x:,.0f}" if pd.notnull(x) else "-"
        def format_percent(x):
            return f"{x*100:.1f}%" if pd.notnull(x) else "-"
        display_df = results_df.copy()
        for col in currency_cols:
            display_df[col] = display_df[col].apply(format_currency)
        display_df[percent_col] = display_df[percent_col].apply(format_percent)
        st.dataframe(
            display_df[
                [
                    "Plan Option Name",
                    "Plan Category",
                    "Network Type",
                    "Total Cost",
                    "Total Premium",
                    "Total OOP",
                    "% of Income",
                ]
            ].reset_index(drop=True),
            use_container_width=True,
        )
        with st.expander("Detailed breakdown for the top 10 plans"):
            st.dataframe(display_df.head(10).reset_index(drop=True), use_container_width=True)
    else:
        st.write("Press **Run comparison** to view your results.")


if __name__ == "__main__":
    main()